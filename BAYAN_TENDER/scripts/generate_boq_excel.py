#!/usr/bin/env python3
"""Generate a proper Excel BOQ file for bid import wizard testing"""
import json
import urllib.request
import os

API_BASE = "http://localhost:5000/api"

def login(email, password):
    data = json.dumps({"email": email, "password": password}).encode()
    req = urllib.request.Request(f"{API_BASE}/auth/login", data=data, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req) as resp:
        result = json.loads(resp.read())
        return result["data"]["accessToken"]

def api_get(token, path):
    req = urllib.request.Request(f"{API_BASE}{path}", headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

# Login
token = login("tendermgr@bayan.ae", "Bayan@2024")

# Get BOQ for TNR-2026-0003
tender_id = "f9d885cc-03eb-435e-a2df-6ff7f1c14041"
boq_data = api_get(token, f"/tenders/{tender_id}/boq")

# Collect all items from all sections
items = []
for section in boq_data.get("data", []):
    for item in section.get("items", []):
        items.append({
            "itemNumber": item["itemNumber"],
            "description": item["description"],
            "quantity": item["quantity"],
            "uom": item["uom"]
        })

print(f"Found {len(items)} BOQ items")
for item in items:
    print(f"  {item['itemNumber']}: {item['description'][:60]}... Qty: {item['quantity']} {item['uom']}")

# Generate Excel file with openpyxl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Priced BOQ"

# Header row - matching what the auto-detection expects
headers = ["Item No.", "Description", "Qty", "UOM", "Unit Rate (AED)", "Amount (AED)", "Currency"]
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 12

# Data rows with realistic prices
unit_prices = {
    "1.1": 850.00,    # Structural concrete
    "1.2": 45.00,     # Steel reinforcement
    "1.3": 380.00,    # Piling
    "1.4": 12.00,     # Formwork
    "1.5": 65.00,     # Waterproofing
    "2.1": 4500000.00, # HVAC LS
    "2.2": 3200000.00, # Electrical LS
    "2.3": 1800000.00, # Plumbing LS
    "2.4": 2100000.00, # Fire protection LS
    "2.5": 850000.00,  # Lifts LS
    "3.1": 185.00,    # Floor tiling
    "3.2": 120.00,    # Suspended ceiling
    "3.3": 35.00,     # Painting
    "3.4": 1250.00,   # Curtain wall
    "3.5": 450.00,    # Kitchen fittings
    "4.1": 280.00,    # Landscaping
    "4.2": 950.00,    # Swimming pool
    "4.3": 22000.00,  # Parking system
    "4.4": 15.00,     # Asphalt
    "4.5": 45.00,     # Boundary wall
}

data_font = Font(name="Arial", size=10)
currency_format = '#,##0.00'

for row_idx, item in enumerate(items, 2):
    item_num = item["itemNumber"]
    qty = item["quantity"]
    unit_price = unit_prices.get(item_num, 500.00)  # Default price if not mapped
    amount = qty * unit_price

    ws.cell(row=row_idx, column=1, value=item_num).font = data_font
    ws.cell(row=row_idx, column=1).border = thin_border

    ws.cell(row=row_idx, column=2, value=item["description"]).font = data_font
    ws.cell(row=row_idx, column=2).border = thin_border
    ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

    ws.cell(row=row_idx, column=3, value=qty).font = data_font
    ws.cell(row=row_idx, column=3).border = thin_border
    ws.cell(row=row_idx, column=3).number_format = '#,##0.00'

    ws.cell(row=row_idx, column=4, value=item["uom"]).font = data_font
    ws.cell(row=row_idx, column=4).border = thin_border

    ws.cell(row=row_idx, column=5, value=unit_price).font = data_font
    ws.cell(row=row_idx, column=5).border = thin_border
    ws.cell(row=row_idx, column=5).number_format = currency_format

    ws.cell(row=row_idx, column=6, value=amount).font = data_font
    ws.cell(row=row_idx, column=6).border = thin_border
    ws.cell(row=row_idx, column=6).number_format = currency_format

    ws.cell(row=row_idx, column=7, value="AED").font = data_font
    ws.cell(row=row_idx, column=7).border = thin_border

# Add total row
total_row = len(items) + 2
total_font = Font(name="Arial", size=11, bold=True)
ws.cell(row=total_row, column=1, value="").font = total_font
ws.cell(row=total_row, column=2, value="TOTAL").font = total_font
ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})").font = total_font
ws.cell(row=total_row, column=6).number_format = currency_format
ws.cell(row=total_row, column=7, value="AED").font = total_font

# Save
output_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
output_path = os.path.join(output_dir, "test-bid-files", "priced-boq-wizard-test.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"\nExcel file saved to: {output_path}")
print(f"File size: {os.path.getsize(output_path)} bytes")
