"""
Bayan Tender Starter Pack Generator
Generates all Excel templates and sample files in the exact format Bayan expects.
Uses a 3-level BOQ hierarchy: Bills -> Items -> Sub-items.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Styling Constants ────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1F3864")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
EDITABLE_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
INSTRUCTION_FONT = Font(name="Calibri", size=10, color="333333")

# Bill-level styling: dark blue background, white bold font
BILL_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
BILL_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

# Item-group styling: light blue background, bold dark font
GROUP_FONT = Font(name="Calibri", bold=True, size=10, color="1F3864")
GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=False)
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── BOQ Hierarchy Data ───────────────────────────────────────────────────────
# 3-level hierarchy: Bills -> Items/Item Groups -> Sub-items
# Realistic commercial building project BOQ (8 bills)
# Uses exact UOM codes from Bayan seed data: m2, sqft, lm, m, nos, kg, ton, ltr, m3, set, lot, LS, day, week, month, %

BOQ_HIERARCHY = [
    # (level, label, description, qty, uom)
    # level: "bill" | "item" | "item_group" | "sub_item"

    # ── BILL 1 ──
    ("bill", "1", "GENERAL REQUIREMENTS", None, None),
    ("item_group", "1.01", "Site Mobilization and Setup", None, None),
    ("sub_item", "a", "Mobilization of plant and equipment", 1, "LS"),
    ("sub_item", "b", "Site office and welfare facilities", 1, "LS"),
    ("sub_item", "c", "Temporary fencing and hoarding", 250, "lm"),
    ("item", "1.02", "Insurance and Bonds", 1, "LS"),
    ("item_group", "1.03", "Health, Safety & Environment", None, None),
    ("sub_item", "a", "HSE plan implementation", 1, "LS"),
    ("sub_item", "b", "PPE and safety equipment", 1, "LS"),
    ("item", "1.04", "Temporary Utilities", 6, "month"),
    ("item", "1.05", "Demobilization and Clearance", 1, "LS"),

    # ── BILL 2 ──
    ("bill", "2", "SUBSTRUCTURE", None, None),
    ("item_group", "2.01", "Excavation Works", None, None),
    ("sub_item", "a", "Excavation in all soil types incl. disposal", 2500, "m3"),
    ("sub_item", "b", "Anti-termite treatment", 800, "m2"),
    ("sub_item", "c", "Backfilling with approved material", 1800, "m3"),
    ("item_group", "2.02", "Foundation Works", None, None),
    ("sub_item", "a", "Plain concrete blinding (C15, 75mm)", 800, "m2"),
    ("sub_item", "b", "Reinforced concrete raft foundation (C40)", 450, "m3"),
    ("sub_item", "c", "Steel reinforcement Grade 60", 54000, "kg"),
    ("sub_item", "d", "Waterproofing membrane to foundation", 800, "m2"),

    # ── BILL 3 ──
    ("bill", "3", "SUPERSTRUCTURE", None, None),
    ("item_group", "3.01", "Concrete Frame", None, None),
    ("sub_item", "a", "RC columns (C40)", 120, "m3"),
    ("sub_item", "b", "RC beams (C40)", 180, "m3"),
    ("sub_item", "c", "RC slabs (C40, 200mm)", 2400, "m2"),
    ("sub_item", "d", "RC core walls (C40)", 85, "m3"),
    ("item_group", "3.02", "Reinforcement and Formwork", None, None),
    ("sub_item", "a", "Steel reinforcement for superstructure", 156000, "kg"),
    ("sub_item", "b", "Formwork to all concrete elements", 4800, "m2"),
    ("item_group", "3.03", "Precast and Steel", None, None),
    ("sub_item", "a", "Precast concrete staircase with landings", 12, "nos"),
    ("sub_item", "b", "Structural steel roof trusses incl. connections", 25, "ton"),

    # ── BILL 4 ──
    ("bill", "4", "ARCHITECTURAL WORKS", None, None),
    ("item_group", "4.01", "Masonry Works", None, None),
    ("sub_item", "a", "External blockwork (200mm AAC)", 3200, "m2"),
    ("sub_item", "b", "Internal blockwork (150mm AAC)", 2800, "m2"),
    ("sub_item", "c", "Internal partitions (100mm AAC)", 1500, "m2"),
    ("item_group", "4.02", "Plastering", None, None),
    ("sub_item", "a", "External cement plaster (20mm, 2 coats)", 3200, "m2"),
    ("sub_item", "b", "Internal cement plaster (15mm)", 8000, "m2"),
    ("item_group", "4.03", "Floor and Wall Finishes", None, None),
    ("sub_item", "a", "Porcelain floor tiles 600x600mm", 2200, "m2"),
    ("sub_item", "b", "Ceramic wall tiles to wet areas", 850, "m2"),
    ("sub_item", "c", "Granite flooring to lobby", 350, "m2"),
    ("sub_item", "d", "Raised access flooring", 1200, "m2"),
    ("item_group", "4.04", "Ceilings and Painting", None, None),
    ("sub_item", "a", "Suspended gypsum board ceiling", 2400, "m2"),
    ("sub_item", "b", "Metal strip ceiling to corridors", 600, "m2"),
    ("sub_item", "c", "Internal painting (3 coats emulsion)", 8000, "m2"),
    ("sub_item", "d", "External painting (3 coats weather-shield)", 3200, "m2"),
    ("item_group", "4.05", "Doors, Windows and Glazing", None, None),
    ("sub_item", "a", "Aluminium curtain wall (double glazed, low-e)", 450, "m2"),
    ("sub_item", "b", "Aluminium windows (double glazed)", 120, "m2"),
    ("sub_item", "c", "Solid core timber doors with ironmongery", 85, "nos"),
    ("sub_item", "d", "Fire rated doors (2hr) with hardware", 24, "nos"),
    ("sub_item", "e", "Automatic sliding glass entrance doors", 4, "nos"),
    ("item_group", "4.06", "Waterproofing and Insulation", None, None),
    ("sub_item", "a", "Roof waterproofing (torch-applied, 2 layers)", 2400, "m2"),
    ("sub_item", "b", "Thermal insulation to roof (50mm XPS)", 2400, "m2"),
    ("sub_item", "c", "Wet area waterproofing (liquid membrane)", 850, "m2"),

    # ── BILL 5 ──
    ("bill", "5", "MECHANICAL WORKS", None, None),
    ("item_group", "5.01", "HVAC System", None, None),
    ("sub_item", "a", "Chilled water AHU units", 8, "nos"),
    ("sub_item", "b", "Fan coil units (various capacities)", 65, "nos"),
    ("sub_item", "c", "VRF outdoor units", 4, "nos"),
    ("sub_item", "d", "Chilled water piping with insulation", 850, "lm"),
    ("sub_item", "e", "GI ductwork incl. fittings", 3200, "kg"),
    ("sub_item", "f", "Air grilles and diffusers", 280, "nos"),
    ("sub_item", "g", "BMS integration for HVAC", 1, "LS"),
    ("item_group", "5.02", "Plumbing and Drainage", None, None),
    ("sub_item", "a", "Hot and cold water piping (PPR)", 1200, "lm"),
    ("sub_item", "b", "Drainage piping (uPVC)", 900, "lm"),
    ("sub_item", "c", "Sanitary fixtures complete", 48, "set"),
    ("sub_item", "d", "Electric water heaters", 12, "nos"),
    ("sub_item", "e", "GRP water storage tank 20m3", 2, "nos"),
    ("sub_item", "f", "Booster pump set (duplex)", 1, "set"),
    ("item_group", "5.03", "Fire Protection", None, None),
    ("sub_item", "a", "Fire sprinkler system complete", 2400, "m2"),
    ("sub_item", "b", "Fire hose reels with cabinets", 16, "nos"),
    ("sub_item", "c", "Portable fire extinguishers", 35, "nos"),
    ("sub_item", "d", "Addressable fire alarm panel", 1, "nos"),
    ("sub_item", "e", "Smoke detectors (addressable)", 180, "nos"),
    ("sub_item", "f", "Fire pump set (electric+diesel+jockey)", 1, "set"),

    # ── BILL 6 ──
    ("bill", "6", "ELECTRICAL WORKS", None, None),
    ("item_group", "6.01", "Power Distribution", None, None),
    ("sub_item", "a", "Main distribution board (4000A)", 1, "nos"),
    ("sub_item", "b", "Sub-main distribution boards", 8, "nos"),
    ("sub_item", "c", "Final distribution boards", 24, "nos"),
    ("sub_item", "d", "Power cables (XLPE/SWA)", 4500, "lm"),
    ("sub_item", "e", "Cable trays and ladders (GI)", 1200, "lm"),
    ("sub_item", "f", "Earthing and lightning protection", 1, "LS"),
    ("item_group", "6.02", "Lighting", None, None),
    ("sub_item", "a", "LED recessed panels 600x600 (40W)", 320, "nos"),
    ("sub_item", "b", "LED downlights (15W adjustable)", 180, "nos"),
    ("sub_item", "c", "Emergency lights with battery", 85, "nos"),
    ("sub_item", "d", "Exit signs (illuminated)", 45, "nos"),
    ("sub_item", "e", "External facade LED lighting", 450, "lm"),
    ("sub_item", "f", "Landscape LED pole lights", 25, "nos"),
    ("item_group", "6.03", "Low Current Systems", None, None),
    ("sub_item", "a", "Structured cabling (Cat6A)", 200, "nos"),
    ("sub_item", "b", "CCTV system (IP, NVR, storage)", 1, "LS"),
    ("sub_item", "c", "Access control system", 1, "LS"),
    ("sub_item", "d", "PA and voice evacuation", 1, "LS"),
    ("sub_item", "e", "VOIP telephone (100 ext)", 1, "LS"),
    ("sub_item", "f", "Standby diesel generator 500KVA", 1, "nos"),
    ("sub_item", "g", "UPS 100KVA online double conversion", 1, "nos"),

    # ── BILL 7 ──
    ("bill", "7", "EXTERNAL WORKS", None, None),
    ("item_group", "7.01", "Roads and Paving", None, None),
    ("sub_item", "a", "Asphalt paving to parking (75mm+150mm)", 3500, "m2"),
    ("sub_item", "b", "Interlocking paver blocks to walkways", 800, "m2"),
    ("sub_item", "c", "Precast concrete kerbs", 650, "lm"),
    ("item_group", "7.02", "Landscaping and Utilities", None, None),
    ("sub_item", "a", "Soft landscaping (trees, shrubs, turf)", 1200, "m2"),
    ("sub_item", "b", "Automatic irrigation system", 1200, "m2"),
    ("sub_item", "c", "External drainage and manholes", 1, "LS"),
    ("item_group", "7.03", "Boundary and Access", None, None),
    ("sub_item", "a", "Precast boundary wall (2.4m high)", 180, "lm"),
    ("sub_item", "b", "Automatic sliding entrance gate (6m)", 2, "nos"),

    # ── BILL 8 ──
    ("bill", "8", "PROVISIONAL SUMS AND DAYWORKS", None, None),
    ("item", "8.01", "Provisional Sums", 1, "LS"),
    ("item", "8.02", "Daywork Allowance", 1, "LS"),
]

# ── Pricing Data for 3 Bidders ──────────────────────────────────────────────
# Hierarchical keys: "bill.item.sub" for sub-items, "bill.item" for standalone items.
# Realistic UAE construction rates.
# Format: {key: (bidder_a_rate, bidder_b_rate, bidder_c_rate)}

BIDDER_PRICES = {
    # Bill 1 - General Requirements
    "1.01.a": (185000, 195000, 175000),
    "1.01.b": (125000, 118000, 135000),
    "1.01.c": (145, 155, 138),
    "1.02": (95000, 88000, 102000),
    "1.03.a": (75000, 82000, 70000),
    "1.03.b": (18000, 16500, 19500),
    "1.04": (18000, 16500, 19500),
    "1.05": (45000, 48000, 42000),
    # Bill 2 - Substructure
    "2.01.a": (65, 72, 60),
    "2.01.b": (28, 25, 32),
    "2.01.c": (55, 62, 50),
    "2.02.a": (85, 90, 78),
    "2.02.b": (1250, 1180, 1320),
    "2.02.c": (4.80, 5.20, 4.50),
    "2.02.d": (95, 88, 102),
    # Bill 3 - Superstructure
    "3.01.a": (1450, 1380, 1520),
    "3.01.b": (1350, 1420, 1280),
    "3.01.c": (285, 298, 270),
    "3.01.d": (1550, 1480, 1620),
    "3.02.a": (4.60, 4.90, 4.35),
    "3.02.b": (120, 135, 112),
    "3.03.a": (15000, 16200, 14500),
    "3.03.b": (12500, 11800, 13200),
    # Bill 4 - Architectural Works
    "4.01.a": (145, 155, 138),
    "4.01.b": (125, 132, 118),
    "4.01.c": (95, 102, 88),
    "4.02.a": (55, 60, 50),
    "4.02.b": (42, 45, 38),
    "4.03.a": (185, 198, 172),
    "4.03.b": (165, 175, 155),
    "4.03.c": (450, 480, 420),
    "4.03.d": (320, 345, 295),
    "4.04.a": (125, 135, 115),
    "4.04.b": (195, 210, 182),
    "4.04.c": (18, 20, 16),
    "4.04.d": (25, 28, 22),
    "4.05.a": (1250, 1180, 1350),
    "4.05.b": (850, 920, 780),
    "4.05.c": (2800, 3050, 2600),
    "4.05.d": (8500, 9200, 7800),
    "4.05.e": (35000, 38000, 32000),
    "4.06.a": (120, 132, 110),
    "4.06.b": (85, 92, 78),
    "4.06.c": (95, 105, 88),
    # Bill 5 - Mechanical Works
    "5.01.a": (85000, 92000, 78000),
    "5.01.b": (4500, 4800, 4200),
    "5.01.c": (125000, 135000, 118000),
    "5.01.d": (185, 198, 170),
    "5.01.e": (32, 35, 28),
    "5.01.f": (280, 310, 255),
    "5.01.g": (95000, 102000, 88000),
    "5.02.a": (85, 92, 78),
    "5.02.b": (65, 72, 58),
    "5.02.c": (4500, 4800, 4200),
    "5.02.d": (3500, 3800, 3200),
    "5.02.e": (45000, 48000, 42000),
    "5.02.f": (65000, 72000, 58000),
    "5.03.a": (55, 60, 50),
    "5.03.b": (3500, 3800, 3200),
    "5.03.c": (450, 490, 410),
    "5.03.d": (85000, 92000, 78000),
    "5.03.e": (350, 380, 320),
    "5.03.f": (250000, 268000, 235000),
    # Bill 6 - Electrical Works
    "6.01.a": (185000, 198000, 172000),
    "6.01.b": (25000, 27000, 23000),
    "6.01.c": (8500, 9200, 7800),
    "6.01.d": (125, 135, 115),
    "6.01.e": (185, 198, 170),
    "6.01.f": (95000, 102000, 88000),
    "6.02.a": (280, 310, 255),
    "6.02.b": (185, 198, 170),
    "6.02.c": (350, 380, 320),
    "6.02.d": (450, 490, 410),
    "6.02.e": (125, 135, 115),
    "6.02.f": (8500, 9200, 7800),
    "6.03.a": (850, 920, 780),
    "6.03.b": (185000, 198000, 172000),
    "6.03.c": (125000, 135000, 115000),
    "6.03.d": (85000, 92000, 78000),
    "6.03.e": (65000, 72000, 58000),
    "6.03.f": (450000, 485000, 420000),
    "6.03.g": (185000, 198000, 172000),
    # Bill 7 - External Works
    "7.01.a": (120, 132, 110),
    "7.01.b": (145, 155, 135),
    "7.01.c": (85, 92, 78),
    "7.02.a": (65, 72, 58),
    "7.02.b": (45, 50, 40),
    "7.02.c": (185000, 198000, 172000),
    "7.03.a": (850, 920, 780),
    "7.03.b": (45000, 48000, 42000),
    # Bill 8 - Provisional Sums and Dayworks
    "8.01": (500000, 520000, 480000),
    "8.02": (150000, 160000, 140000),
}

BIDDER_NAMES = {
    "a": "Al Futtaim Construction LLC",
    "b": "Emirates Building Systems Co.",
    "c": "Gulf Contracting Company WLL",
}


# ── Helper: build price key for each row ─────────────────────────────────────
def _build_price_key(level, label, current_bill, current_item):
    """Build the BIDDER_PRICES lookup key for a given hierarchy row."""
    if level == "item":
        # Standalone item: key is "bill.item" e.g. "1.02"
        return label
    if level == "sub_item":
        # Sub-item: key is "bill.item_group.sub" e.g. "1.01.a"
        return f"{current_item}.{label}"
    return None


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_bill_row(ws, row, max_col):
    """Apply bill-level styling: dark blue background, white bold font."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BILL_FONT
        cell.fill = BILL_FILL
        cell.border = THIN_BORDER


def style_group_row(ws, row, max_col):
    """Apply item-group styling: light blue background, bold dark font."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = GROUP_FONT
        cell.fill = GROUP_FILL
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, is_locked=False):
    """Apply data row styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if is_locked:
            cell.fill = LOCKED_FILL


def add_instructions_sheet(wb, is_bidder=False, bidder_name=None):
    """Add an Instructions sheet to the workbook (as last sheet, not first)."""
    ws = wb.create_sheet("Instructions")

    ws.column_dimensions["A"].width = 80
    ws.sheet_properties.tabColor = "2F5496"

    row = 1
    ws.cell(row=row, column=1, value="BAYAN TENDER MANAGEMENT SYSTEM").font = TITLE_FONT
    row += 1
    if is_bidder:
        ws.cell(row=row, column=1, value=f"PRICED BILL OF QUANTITIES - {bidder_name}").font = SUBTITLE_FONT
    else:
        ws.cell(row=row, column=1, value="BILL OF QUANTITIES").font = SUBTITLE_FONT
    row += 2

    instructions = [
        "IMPORTANT INSTRUCTIONS:",
        "",
        "1. Do NOT modify the column headers in the 'BOQ' sheet.",
        "2. The system auto-detects columns by header name.",
        '3. Required headers: "Bill No.", "Item No.", "Sub-Item", "Description", "Qty", "Unit"',
        "",
        "BOQ HIERARCHY:",
        "  This BOQ uses a 3-level hierarchy:",
        "  - BILL (Level 1): Major trade sections (dark blue rows)",
        "  - ITEM / ITEM GROUP (Level 2): Work categories within a bill (light blue rows for groups, white for standalone items)",
        "  - SUB-ITEM (Level 3): Individual line items within a group (indented, lettered a, b, c...)",
        "",
        "  Example:",
        "    Bill 1 - GENERAL REQUIREMENTS",
        "      1.01 - Site Mobilization and Setup (item group)",
        "        a) Mobilization of plant and equipment",
        "        b) Site office and welfare facilities",
        "      1.02 - Insurance and Bonds (standalone item with qty/unit)",
        "",
    ]

    if is_bidder:
        instructions += [
            "FOR BIDDERS:",
            '4. Enter your Unit Rate in the "Unit Rate" column (yellow-highlighted cells only).',
            '5. The "Amount" column shows Qty x Unit Rate - do NOT modify.',
            "6. Only leaf-level rows (standalone items and sub-items) have editable rate cells.",
            "7. Group rows show subtotals, bill rows show bill totals.",
            "8. All prices must be in AED (UAE Dirhams).",
            "9. Zero-value items will be flagged for confirmation.",
            "10. Negative values are NOT allowed.",
            "",
            "FORMULA VALIDATION:",
            "- Bayan validates: Amount = Qty x Unit Rate (1% tolerance).",
            "- Do NOT manually override the Amount values.",
            "",
        ]
    else:
        instructions += [
            "FOR TENDER MANAGERS:",
            "4. Bill rows have Bill No. and Description only (dark blue).",
            "5. Item group rows have Item No. and Description only (light blue).",
            "6. Standalone items have Item No., Description, Qty, and Unit.",
            "7. Sub-items have Sub-Item letter, Description, Qty, and Unit.",
            "8. Bayan auto-detects the hierarchy from row styling and column presence.",
            '9. Valid UOM codes: m2, sqft, lm, m, nos, kg, ton, ltr, m3, set, lot, LS, day, week, month, %',
            "",
        ]

    instructions += [
        "SUPPORTED FILE FORMATS:",
        "- .xlsx (recommended)",
        "- .xls",
        "- .csv",
        "",
        "COLUMN AUTO-DETECTION KEYWORDS:",
        '- Bill Number: "Bill No.", "Bill #", "Bill"',
        '- Item Number: "Item No.", "Item #", "Ref", "Code"',
        '- Sub-Item: "Sub-Item", "Sub Item", "Sub", "Detail"',
        '- Description: "Description", "Particulars", "Detail", "Scope"',
        '- Quantity: "Qty", "Quantity", "Est. Qty"',
        '- UOM: "Unit", "UOM", "Unit of Measurement"',
        '- Unit Rate: "Unit Rate", "Rate", "Unit Price", "Price"',
        '- Amount: "Amount", "Total", "Total Amount", "Value"',
        '- Notes: "Notes", "Remarks", "Comments"',
    ]

    for line in instructions:
        ws.cell(row=row, column=1, value=line).font = INSTRUCTION_FONT
        row += 1


def create_empty_boq_template():
    """Create the empty BOQ template with 7-column headers."""
    wb = Workbook()

    # BOQ Sheet (first/active sheet)
    ws = wb.active
    ws.title = "BOQ"

    # Instructions sheet (second)
    add_instructions_sheet(wb)

    headers = ["Bill No.", "Item No.", "Sub-Item", "Description", "Qty", "Unit", "Notes"]
    col_widths = [10, 10, 10, 55, 12, 10, 25]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    style_header_row(ws, 1, len(headers))

    # Add a few empty formatted rows to guide the user
    for row in range(2, 12):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # Freeze header row
    ws.freeze_panes = "A2"

    path = os.path.join(BASE_DIR, "01_BOQ_Template_Empty.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


def create_filled_boq_no_prices():
    """Create a filled BOQ with hierarchical items but no prices (master BOQ).

    Columns: Bill No., Item No., Sub-Item, Description, Qty, Unit
    Styling:
      - Bill rows: Dark blue background (#1F3864), white bold font
      - Item GROUP rows: Light blue (#D6E4F0), bold, no Qty/Unit
      - Item STANDALONE: White, normal data, has Qty/Unit
      - Sub-item: White, indented description ("  a) ..."), has Qty/Unit
    """
    wb = Workbook()

    # BOQ Sheet (first/active sheet)
    ws = wb.active
    ws.title = "BOQ"

    # Instructions sheet (second)
    add_instructions_sheet(wb)

    headers = ["Bill No.", "Item No.", "Sub-Item", "Description", "Qty", "Unit"]
    col_widths = [10, 10, 10, 60, 12, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    style_header_row(ws, 1, len(headers))

    row = 2
    current_bill = None

    for level, label, desc, qty, uom in BOQ_HIERARCHY:
        if level == "bill":
            current_bill = label
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=4, value=desc)
            style_bill_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN

        elif level == "item_group":
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=4, value=desc)
            style_group_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN

        elif level == "item":
            # Standalone item — has Qty and Unit
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=4, value=desc)
            ws.cell(row=row, column=5, value=qty)
            ws.cell(row=row, column=6, value=uom)
            style_data_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN
            ws.cell(row=row, column=5).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = CENTER_ALIGN

        elif level == "sub_item":
            # Sub-item — indented description, has Qty and Unit
            ws.cell(row=row, column=3, value=label)
            ws.cell(row=row, column=4, value=f"  {label}) {desc}")
            ws.cell(row=row, column=5, value=qty)
            ws.cell(row=row, column=6, value=uom)
            style_data_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN
            ws.cell(row=row, column=5).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = CENTER_ALIGN

        row += 1

    ws.freeze_panes = "A2"

    path = os.path.join(BASE_DIR, "02_BOQ_Master_No_Prices.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


def create_priced_boq(bidder_key):
    """Create a priced BOQ for a specific bidder with hierarchical subtotals.

    Columns: Bill No., Item No., Sub-Item, Description, Qty, Unit, Unit Rate, Amount
    - Leaf rows (item/sub_item) have editable Unit Rate (yellow) and pre-calculated Amount
    - Item group rows show subtotal in Amount column
    - Bill rows show bill total in Amount column
    - Grand total at bottom with dark blue styling
    """
    bidder_name = BIDDER_NAMES[bidder_key]
    price_index = {"a": 0, "b": 1, "c": 2}[bidder_key]

    wb = Workbook()

    # BOQ Sheet (first/active sheet)
    ws = wb.active
    ws.title = "BOQ"

    # Instructions sheet (second)
    add_instructions_sheet(wb, is_bidder=True, bidder_name=bidder_name)

    headers = ["Bill No.", "Item No.", "Sub-Item", "Description", "Qty", "Unit", "Unit Rate", "Amount"]
    col_widths = [10, 10, 10, 60, 12, 10, 18, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    style_header_row(ws, 1, len(headers))

    # Pre-calculate all amounts and subtotals
    # We need two passes: first calculate amounts, then write with subtotals
    # Parse hierarchy into a structure that allows subtotaling
    current_bill = None
    current_item = None
    bill_totals = {}       # bill_label -> total amount
    group_totals = {}      # item_label -> total amount (for item_groups)
    item_amounts = {}      # price_key -> amount

    for level, label, desc, qty, uom in BOQ_HIERARCHY:
        if level == "bill":
            current_bill = label
            if current_bill not in bill_totals:
                bill_totals[current_bill] = 0.0
        elif level == "item_group":
            current_item = label
            if current_item not in group_totals:
                group_totals[current_item] = 0.0
        elif level == "item":
            # Standalone item
            price_key = label
            prices = BIDDER_PRICES.get(price_key)
            if prices and qty is not None:
                unit_rate = prices[price_index]
                amount = round(qty * unit_rate, 2)
                item_amounts[price_key] = amount
                bill_totals[current_bill] = bill_totals.get(current_bill, 0.0) + amount
        elif level == "sub_item":
            # Sub-item under current_item group
            price_key = f"{current_item}.{label}"
            prices = BIDDER_PRICES.get(price_key)
            if prices and qty is not None:
                unit_rate = prices[price_index]
                amount = round(qty * unit_rate, 2)
                item_amounts[price_key] = amount
                group_totals[current_item] = group_totals.get(current_item, 0.0) + amount
                bill_totals[current_bill] = bill_totals.get(current_bill, 0.0) + amount

    # Now write to Excel
    row = 2
    current_bill = None
    current_item = None
    grand_total = 0.0

    for level, label, desc, qty, uom in BOQ_HIERARCHY:
        if level == "bill":
            current_bill = label
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=4, value=desc)
            # Bill total in Amount column
            bt = round(bill_totals.get(label, 0.0), 2)
            ws.cell(row=row, column=8, value=bt)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
            grand_total += bt
            style_bill_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN

        elif level == "item_group":
            current_item = label
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=4, value=desc)
            # Group subtotal in Amount column
            gt = round(group_totals.get(label, 0.0), 2)
            ws.cell(row=row, column=8, value=gt)
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
            style_group_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN

        elif level == "item":
            # Standalone item
            price_key = label
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=4, value=desc)
            ws.cell(row=row, column=5, value=qty)
            ws.cell(row=row, column=6, value=uom)

            prices = BIDDER_PRICES.get(price_key)
            if prices:
                unit_rate = prices[price_index]
                amount = item_amounts.get(price_key, 0.0)
                ws.cell(row=row, column=7, value=unit_rate)
                ws.cell(row=row, column=8, value=amount)

            style_data_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN
            ws.cell(row=row, column=5).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = CENTER_ALIGN
            ws.cell(row=row, column=7).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            # Editable rate cell (yellow)
            ws.cell(row=row, column=7).fill = EDITABLE_FILL

        elif level == "sub_item":
            # Sub-item
            price_key = f"{current_item}.{label}"
            ws.cell(row=row, column=3, value=label)
            ws.cell(row=row, column=4, value=f"  {label}) {desc}")
            ws.cell(row=row, column=5, value=qty)
            ws.cell(row=row, column=6, value=uom)

            prices = BIDDER_PRICES.get(price_key)
            if prices:
                unit_rate = prices[price_index]
                amount = item_amounts.get(price_key, 0.0)
                ws.cell(row=row, column=7, value=unit_rate)
                ws.cell(row=row, column=8, value=amount)

            style_data_row(ws, row, len(headers))
            ws.cell(row=row, column=4).alignment = LEFT_ALIGN
            ws.cell(row=row, column=5).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).alignment = CENTER_ALIGN
            ws.cell(row=row, column=7).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8).alignment = RIGHT_ALIGN
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            # Editable rate cell (yellow)
            ws.cell(row=row, column=7).fill = EDITABLE_FILL

        row += 1

    # Grand Total row
    row += 1
    ws.cell(row=row, column=4, value="GRAND TOTAL (AED)").font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=row, column=4).alignment = LEFT_ALIGN

    ws.cell(row=row, column=8, value=round(grand_total, 2))
    ws.cell(row=row, column=8).font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=8).alignment = RIGHT_ALIGN

    # Style the total row: dark blue
    total_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    letter = bidder_key.upper()
    filename = f"03_Priced_BOQ_Bidder_{letter}_{bidder_name.replace(' ', '_').replace('.', '')}.xlsx"
    path = os.path.join(BASE_DIR, filename)
    wb.save(path)
    print(f"  Created: {path}")


def create_evaluation_criteria():
    """Create evaluation criteria template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation Criteria"

    # Title
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="TENDER EVALUATION CRITERIA").font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value="Commercial Building Project - Abu Dhabi").font = SUBTITLE_FONT

    # Technical Criteria
    row = 4
    ws.cell(row=row, column=1, value="TECHNICAL EVALUATION (40%)").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    row += 1

    headers = ["#", "Criterion", "Weight (%)", "Guidance Notes"]
    col_widths = [5, 30, 12, 50]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    style_header_row(ws, row, len(headers))

    criteria = [
        (1, "Compliance", 20, "Assessment of bid completeness, adherence to tender requirements, and compliance with specifications. Check all mandatory documents are submitted."),
        (2, "Methodology", 20, "Evaluate the proposed construction methodology, work sequence, quality of technical approach, and innovation in execution methods."),
        (3, "Team CVs", 15, "Review qualifications, experience, and certifications of key personnel. Minimum 10 years for PM, 7 years for site engineers."),
        (4, "Program", 15, "Assess the construction schedule, critical path analysis, milestone planning, and resource allocation. Must show weather contingency."),
        (5, "QA/QC", 15, "Evaluate the quality assurance and quality control plan, inspection procedures, testing protocols, and material approval process."),
        (6, "HSE", 15, "Review the Health, Safety & Environment plan, emergency procedures, incident reporting, and compliance with UAE regulations."),
    ]

    for num, name, weight, notes in criteria:
        row += 1
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=weight)
        ws.cell(row=row, column=4, value=notes)
        style_data_row(ws, row, len(headers))
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=4).alignment = WRAP_ALIGN
        ws.row_dimensions[row].height = 45

    row += 1
    ws.cell(row=row, column=2, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=3, value=100).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=3).alignment = CENTER_ALIGN

    # Commercial Section
    row += 2
    ws.cell(row=row, column=1, value="COMMERCIAL EVALUATION (60%)").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    row += 1
    ws.cell(row=row, column=1, value="Commercial scoring is automatic based on lowest bidder normalization formula:").font = INSTRUCTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Score = (Lowest Bid / Bidder's Bid) x 100").font = Font(name="Calibri", bold=True, italic=True, size=10)

    # Weights Summary
    row += 2
    ws.cell(row=row, column=1, value="COMBINED SCORING FORMULA").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    row += 1
    ws.cell(row=row, column=1, value="Final Score = (Technical Score x 0.40) + (Commercial Score x 0.60)").font = Font(name="Calibri", bold=True, italic=True, size=10)

    path = os.path.join(BASE_DIR, "04_Evaluation_Criteria.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


def create_bidder_list():
    """Create a bidder list for invitation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bidders"

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="INVITED BIDDERS LIST").font = TITLE_FONT

    row = 3
    headers = ["#", "Company Name", "Contact Person", "Email", "Phone", "Qualification Status"]
    col_widths = [5, 35, 25, 30, 18, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    style_header_row(ws, row, len(headers))

    bidders = [
        (1, "Al Futtaim Construction LLC", "Ahmed Al Futtaim", "ahmed@alfuttaim-const.ae", "+971 4 555 0101", "Qualified"),
        (2, "Emirates Building Systems Co.", "Mohamed Al Emirati", "mohamed@ebs-uae.ae", "+971 2 555 0202", "Qualified"),
        (3, "Gulf Contracting Company WLL", "Khalid Al Mansoori", "khalid@gulfcontracting.ae", "+971 6 555 0303", "Qualified"),
    ]

    for num, name, contact, email, phone, status in bidders:
        row += 1
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=contact)
        ws.cell(row=row, column=4, value=email)
        ws.cell(row=row, column=5, value=phone)
        ws.cell(row=row, column=6, value=status)
        style_data_row(ws, row, len(headers))
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=6).alignment = CENTER_ALIGN

    path = os.path.join(BASE_DIR, "05_Bidder_List.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


def create_tender_details():
    """Create tender details reference document."""
    content = """# Tender Details - Ready for Bayan Input

## Basic Information
| Field | Value |
|-------|-------|
| **Title** | Construction of Commercial Office Building - Plot C7, Al Reem Island |
| **Description** | Design and Build of a 6-storey commercial office building including basement parking, ground floor retail, and 5 floors of office space. Total BUA approximately 12,000 sqm. |
| **Tender Type** | Selective |
| **Base Currency** | AED |
| **Estimated Value** | 45,000,000 |
| **Bid Validity (Days)** | 90 |
| **Pricing Level** | SubItem |

## Key Dates
| Date Field | Value |
|------------|-------|
| **Issue Date** | 2026-03-01 |
| **Clarification Deadline** | 2026-03-15 |
| **Submission Deadline** | 2026-04-01 |
| **Opening Date** | 2026-04-02 |

## BOQ Structure
| Level | Count | Description |
|-------|-------|-------------|
| **Bills** | 8 | Major trade sections |
| **Items / Item Groups** | ~26 | Work categories within bills |
| **Sub-Items** | ~80 | Individual priced line items |

## Evaluation Weights
| Component | Weight |
|-----------|--------|
| **Technical** | 40% |
| **Commercial** | 60% |

## Evaluation Criteria
| Criterion | Weight | Guidance |
|-----------|--------|----------|
| Compliance | 20% | Bid completeness and adherence to requirements |
| Methodology | 20% | Technical approach and construction methods |
| Team CVs | 15% | Key personnel qualifications and experience |
| Program | 15% | Construction schedule and resource planning |
| QA/QC | 15% | Quality management systems and procedures |
| HSE | 15% | Health, safety & environment plans |

## Invited Bidders
1. **Al Futtaim Construction LLC** - ahmed@alfuttaim-const.ae
2. **Emirates Building Systems Co.** - mohamed@ebs-uae.ae
3. **Gulf Contracting Company WLL** - khalid@gulfcontracting.ae

## Required Bid Documents
| Document | Type | Required |
|----------|------|----------|
| Priced BOQ | Commercial | Yes |
| Methodology | Technical | Yes |
| Team CVs | Technical | Yes |
| Program | Technical | Yes |
| HSE Plan | Technical | Yes |
| Supporting Docs | Other | Optional |

## Presentation Flow (E2E Demo)
1. **Create Tender** - Enter all details above
2. **Upload Master BOQ** - Use `02_BOQ_Master_No_Prices.xlsx`
3. **Set Evaluation Criteria** - Use `04_Evaluation_Criteria.xlsx`
4. **Invite Bidders** - Use `05_Bidder_List.xlsx`
5. **Publish Tender** - Transition to Active status
6. **Import Bid A** - Upload `03_Priced_BOQ_Bidder_A_*.xlsx`
7. **Import Bid B** - Upload `03_Priced_BOQ_Bidder_B_*.xlsx`
8. **Import Bid C** - Upload `03_Priced_BOQ_Bidder_C_*.xlsx`
9. **Technical Evaluation** - Score each bidder
10. **Commercial Evaluation** - Auto-calculated from BOQ
11. **Combined Scoring** - View final rankings
12. **Award** - Select winner and initiate approval
"""

    path = os.path.join(BASE_DIR, "06_Tender_Details.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Created: {path}")


def create_uom_reference():
    """Create UOM reference sheet matching Bayan's seeded values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UOM Reference"

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="BAYAN - VALID UNIT OF MEASUREMENT CODES").font = TITLE_FONT

    row = 3
    headers = ["Code", "Name", "Category", "Base Unit", "Conversion Factor"]
    col_widths = [10, 20, 15, 12, 18]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=row, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    style_header_row(ws, row, len(headers))

    uoms = [
        ("m2", "Square Meter", "Area", "-", 1.0),
        ("sqft", "Square Foot", "Area", "m2", 0.092903),
        ("lm", "Linear Meter", "Length", "-", 1.0),
        ("m", "Meter", "Length", "lm", 1.0),
        ("nos", "Numbers", "Count", "-", 1.0),
        ("kg", "Kilogram", "Weight", "-", 1.0),
        ("ton", "Metric Ton", "Weight", "kg", 1000.0),
        ("ltr", "Liter", "Volume", "m3", 0.001),
        ("m3", "Cubic Meter", "Volume", "-", 1.0),
        ("set", "Set", "Lump", "-", 1.0),
        ("lot", "Lot", "Lump", "-", 1.0),
        ("LS", "Lump Sum", "Lump", "-", 1.0),
        ("day", "Day", "Time", "-", 1.0),
        ("week", "Week", "Time", "day", 7.0),
        ("month", "Month", "Time", "day", 30.0),
        ("%", "Percentage", "Percentage", "-", 1.0),
    ]

    for code, name, cat, base, cf in uoms:
        row += 1
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=base)
        ws.cell(row=row, column=5, value=cf)
        style_data_row(ws, row, len(headers))
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5).alignment = RIGHT_ALIGN

    path = os.path.join(BASE_DIR, "07_UOM_Reference.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


def create_presentation_checklist():
    """Create a presentation checklist."""
    content = """# Bayan E2E Presentation Checklist

## Pre-Presentation Setup
- [ ] Bayan application is running and accessible
- [ ] All starter pack files are downloaded and accessible
- [ ] Test user accounts are ready (TenderManager, CommercialAnalyst, etc.)
- [ ] Browser is set up (Chrome recommended)

## Demo Flow

### Phase 1: Tender Creation (5 min)
- [ ] Log in as Tender Manager
- [ ] Click "Create Tender"
- [ ] Fill in details from `06_Tender_Details.md`
  - Title: Construction of Commercial Office Building - Plot C7, Al Reem Island
  - Type: Selective
  - Currency: AED
  - Estimated Value: 45,000,000
  - Dates: Issue=Mar 1, Clarification=Mar 15, Submission=Apr 1, Opening=Apr 2
  - Tech Weight: 40%, Commercial Weight: 60%
- [ ] Save as Draft

### Phase 2: BOQ Upload (3 min)
- [ ] Navigate to BOQ tab
- [ ] Click "Import BOQ"
- [ ] Upload `02_BOQ_Master_No_Prices.xlsx`
- [ ] Review auto-detected column mapping
  - Bill No. -> BillNumber
  - Item No. -> ItemNumber
  - Sub-Item -> SubItem
  - Description -> Description
  - Qty -> Quantity
  - Unit -> Uom
- [ ] Select Pricing Level: **SubItem** (3-level hierarchy)
- [ ] Validate (should show 0 errors)
- [ ] Import and verify hierarchy: 8 bills, ~26 items, ~80 sub-items
- [ ] Show visual hierarchy in BOQ viewer (dark blue bills, light blue groups, indented sub-items)

### Phase 3: Evaluation Criteria (2 min)
- [ ] Navigate to Evaluation tab
- [ ] Add criteria matching `04_Evaluation_Criteria.xlsx`:
  - Compliance: 20%
  - Methodology: 20%
  - Team CVs: 15%
  - Program: 15%
  - QA/QC: 15%
  - HSE: 15%
- [ ] Total = 100% (validation passes)

### Phase 4: Invite Bidders (2 min)
- [ ] Navigate to Bidders tab
- [ ] Add 3 bidders from `05_Bidder_List.xlsx`
- [ ] Show invitation status

### Phase 5: Publish Tender (1 min)
- [ ] Click "Publish" to move from Draft -> Active
- [ ] Show tender is now live

### Phase 6: Bid Submission (5 min)
- [ ] Navigate to Bids/Submissions section
- [ ] Import Bidder A: Upload `03_Priced_BOQ_Bidder_A_*.xlsx`
  - Map columns: Bill No., Item No., Sub-Item, Description, Qty, Unit, Unit Rate, Amount
  - Validate: should show 0 errors, formula check passes
  - Import
- [ ] Import Bidder B: Upload `03_Priced_BOQ_Bidder_B_*.xlsx`
  - Same process
- [ ] Import Bidder C: Upload `03_Priced_BOQ_Bidder_C_*.xlsx`
  - Same process
- [ ] Show all 3 bids imported successfully
- [ ] Verify subtotals per bill and grand total

### Phase 7: Technical Evaluation (3 min)
- [ ] Close tender for evaluation (Active -> Evaluation)
- [ ] Navigate to Technical Evaluation
- [ ] Score each bidder on all criteria (sample scores)
- [ ] Lock technical scores

### Phase 8: Commercial Evaluation (2 min)
- [ ] Navigate to Commercial Evaluation
- [ ] Show auto-calculated commercial scores
- [ ] Show price comparison and breakdown by bill

### Phase 9: Combined Scoring (2 min)
- [ ] Navigate to Combined Scorecard
- [ ] Show final rankings
- [ ] Show weighted formula: (Tech x 0.40) + (Commercial x 0.60)

### Phase 10: Award (2 min)
- [ ] Select winning bidder
- [ ] Initiate approval workflow
- [ ] Show award pack generation

## Key Talking Points
1. **3-Level Hierarchy**: Bills > Items > Sub-items with visual differentiation
2. **Auto-detection**: Bayan auto-maps Excel columns by header names
3. **Subtotals**: Group subtotals, bill totals, and grand total auto-calculated
4. **Validation**: Built-in formula validation (Qty x Rate = Amount, 1% tolerance)
5. **Section hierarchy**: Auto-creates from bill/item/sub-item structure
6. **Fair evaluation**: Blind mode separates technical from commercial
7. **Audit trail**: Full activity log and document versioning
8. **Multi-approval**: Sequential approval workflow for awards

## Troubleshooting
| Issue | Solution |
|-------|----------|
| Column not detected | Check header matches keywords exactly |
| Validation errors | Ensure no empty Item No. or Description |
| Formula mismatch | Verify Amount = Qty x Unit Rate (1% tolerance) |
| UOM warning | Use codes from `07_UOM_Reference.xlsx` |
| Import fails | Ensure .xlsx format, max 50MB |
| Hierarchy not shown | Ensure Pricing Level is set to SubItem |
"""

    path = os.path.join(BASE_DIR, "08_Presentation_Checklist.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Created: {path}")


# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating Bayan Tender Starter Pack...")
    print("=" * 50)

    create_empty_boq_template()
    create_filled_boq_no_prices()
    create_priced_boq("a")
    create_priced_boq("b")
    create_priced_boq("c")
    create_evaluation_criteria()
    create_bidder_list()
    create_tender_details()
    create_uom_reference()
    create_presentation_checklist()

    print("=" * 50)
    print("Starter pack generation complete!")
    print(f"All files saved to: {BASE_DIR}")
