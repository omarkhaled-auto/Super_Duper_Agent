"""Generate a clean flat priced BOQ xlsx from tender BOQ items — no section headers or subtotals."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

BOQ_ITEMS = [
    ("E-001", "Supply and install main LV switchboard panel (MCCB 3200A)", 1.0, "No.", 285000.00),
    ("E-002", "Supply and install sub-distribution boards (SDB) per floor", 24.0, "No.", 12500.00),
    ("E-003", "Supply and install XLPE copper power cables 4x240mm2", 2500.0, "LM", 485.00),
    ("E-004", "Supply and install XLPE copper power cables 4x95mm2", 4800.0, "LM", 195.00),
    ("E-005", "Supply and install lighting fixtures - LED panel 60x60cm", 1200.0, "No.", 185.00),
    ("E-006", "Supply and install lighting fixtures - LED downlight 8 inch", 850.0, "No.", 145.00),
    ("E-007", "Supply and install emergency lighting with battery backup", 320.0, "No.", 320.00),
    ("E-008", "Supply and install cable trays - hot dip galvanized 600mm", 3200.0, "LM", 125.00),
    ("E-009", "Supply and install cable trays - hot dip galvanized 300mm", 2100.0, "LM", 85.00),
    ("E-010", "Supply and install power sockets 13A twin switched", 960.0, "No.", 65.00),
    ("E-011", "Supply and install 20A outlet for AC units", 96.0, "No.", 95.00),
    ("E-012", "Supply and install earthing and lightning protection system", 1.0, "Lot", 175000.00),
    ("F-001", "Supply and install fire pump set - electric main pump", 1.0, "Set", 245000.00),
    ("F-002", "Supply and install fire pump set - diesel standby pump", 1.0, "Set", 320000.00),
    ("F-003", "Supply and install fire pump set - jockey pump", 1.0, "Set", 45000.00),
    ("F-004", "Supply and install wet riser piping - SCH40 150mm", 800.0, "LM", 285.00),
    ("F-005", "Supply and install wet riser piping - SCH40 65mm", 2400.0, "LM", 125.00),
    ("F-006", "Supply and install sprinkler heads - pendant K5.6", 2400.0, "No.", 48.00),
    ("F-007", "Supply and install fire hose cabinets - recessed", 48.0, "No.", 2800.00),
    ("F-008", "Supply and install fire alarm control panel - addressable", 1.0, "No.", 185000.00),
    ("F-009", "Supply and install smoke detectors - addressable optical", 960.0, "No.", 165.00),
    ("F-010", "Supply and install manual call points - addressable", 96.0, "No.", 195.00),
    ("F-011", "Supply and install fire alarm sounders with beacon", 192.0, "No.", 225.00),
    ("F-012", "Supply and install fire rated cabling 2C 1.5mm2", 12000.0, "LM", 18.50),
    ("H-001", "Supply and install chilled water AHU - 20,000 CFM", 4.0, "No.", 145000.00),
    ("H-002", "Supply and install chilled water AHU - 10,000 CFM", 8.0, "No.", 82000.00),
    ("H-003", "Supply and install fan coil units - ceiling concealed 2TR", 120.0, "No.", 3200.00),
    ("H-004", "Supply and install fan coil units - ceiling concealed 1TR", 80.0, "No.", 2400.00),
    ("H-005", "Supply and install chilled water piping - insulated 150mm", 1800.0, "LM", 245.00),
    ("H-006", "Supply and install chilled water piping - insulated 100mm", 2400.0, "LM", 175.00),
    ("H-007", "Supply and install chilled water piping - insulated 50mm", 3600.0, "LM", 95.00),
    ("H-008", "Supply and install GI ductwork - rectangular", 8500.0, "Kg", 32.00),
    ("H-009", "Supply and install flexible duct connections", 400.0, "No.", 85.00),
    ("H-010", "Supply and install supply air diffusers - 600x600mm", 480.0, "No.", 120.00),
    ("H-011", "Supply and install return air grilles - 600x300mm", 480.0, "No.", 95.00),
    ("H-012", "Supply and install BMS controllers and sensors", 1.0, "Lot", 450000.00),
    ("P-001", "Supply and install domestic water pumps - booster set", 2.0, "Set", 65000.00),
    ("P-002", "Supply and install PPR hot water piping 63mm", 1200.0, "LM", 75.00),
    ("P-003", "Supply and install PPR cold water piping 75mm", 1800.0, "LM", 85.00),
    ("P-004", "Supply and install PPR cold water piping 32mm", 3200.0, "LM", 42.00),
    ("P-005", "Supply and install uPVC drainage piping 110mm", 2800.0, "LM", 65.00),
    ("P-006", "Supply and install uPVC drainage piping 50mm", 1600.0, "LM", 38.00),
    ("P-007", "Supply and install floor drains - stainless steel 150mm", 320.0, "No.", 145.00),
    ("P-008", "Supply and install water heaters - 100L electric", 24.0, "No.", 2200.00),
    ("P-009", "Supply and install sanitary fixtures - WC wall hung", 96.0, "No.", 1850.00),
    ("P-010", "Supply and install sanitary fixtures - wash basin counter top", 96.0, "No.", 1450.00),
    ("P-011", "Supply and install kitchen sink - double bowl SS", 24.0, "No.", 1800.00),
    ("P-012", "Supply and install GRP water storage tank 50,000L", 2.0, "No.", 185000.00),
]

OUTPUT_PATH = r"C:\Users\Omar Khaled\OneDrive\Desktop\claude-agent-team\BAYAN_TENDER\PricedBOQ_ABCConstruction.xlsx"

def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Priced BOQ"

    header_font = Font(name="Calibri", bold=True, size=11)
    data_font = Font(name="Calibri", size=10)
    money_fmt = '#,##0.00'
    qty_fmt = '#,##0.0000'
    thin_border = Border(
        left=Side(style='thin', color='D4D4D8'),
        right=Side(style='thin', color='D4D4D8'),
        top=Side(style='thin', color='D4D4D8'),
        bottom=Side(style='thin', color='D4D4D8'),
    )

    # Headers in row 1
    headers = ["Item No.", "Description", "Qty", "UOM", "Unit Rate (AED)", "Amount (AED)"]
    col_widths = [12, 55, 12, 10, 18, 20]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Data rows starting at row 2
    for row_idx, (item_no, desc, qty, uom, rate) in enumerate(BOQ_ITEMS, 2):
        amount = round(qty * rate, 2)

        ws.cell(row=row_idx, column=1, value=item_no).font = data_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=desc).font = data_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=2).border = thin_border

        c = ws.cell(row=row_idx, column=3, value=qty)
        c.font = data_font
        c.number_format = qty_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

        ws.cell(row=row_idx, column=4, value=uom).font = data_font
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=4).border = thin_border

        c = ws.cell(row=row_idx, column=5, value=rate)
        c.font = data_font
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

        c = ws.cell(row=row_idx, column=6, value=amount)
        c.font = data_font
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = thin_border

    wb.save(OUTPUT_PATH)
    grand_total = sum(round(q * r, 2) for _, _, q, _, r in BOQ_ITEMS)
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Grand Total: AED {grand_total:,.2f}")
    print(f"Items: {len(BOQ_ITEMS)} (flat table, no section headers)")

    import os
    print(f"File size: {os.path.getsize(OUTPUT_PATH)} bytes")

if __name__ == "__main__":
    generate()
